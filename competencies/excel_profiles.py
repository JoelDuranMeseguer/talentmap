import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from competencies.models import Competency, RoleCompetencyRequirement
from people.models import Role

HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")


def build_role_profile_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "PerfilIdeal"

    roles = list(Role.objects.select_related("department").order_by("department__name", "name"))
    competencies = list(Competency.objects.order_by("name"))
    reqs = {
        (r.competency_id, r.role_id): r
        for r in RoleCompetencyRequirement.objects.select_related("competency", "role")
    }

    ws.cell(row=1, column=1, value="Competencia")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).fill = HEADER_FILL

    for idx, role in enumerate(roles, start=2):
        ws.cell(row=1, column=idx, value=role.name)
        ws.cell(row=1, column=idx).font = Font(bold=True)
        ws.cell(row=1, column=idx).fill = HEADER_FILL
        ws.cell(row=2, column=idx, value=f"{role.department.name}")
        ws.cell(row=2, column=idx).alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value="Nivel requerido (1=Básico, 2=Avanzado, 3=Experto)")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    for row_idx, competency in enumerate(competencies, start=3):
        ws.cell(row=row_idx, column=1, value=competency.name)
        for col_idx, role in enumerate(roles, start=2):
            req = reqs.get((competency.id, role.id))
            if req:
                ws.cell(row=row_idx, column=col_idx, value=int(req.required_level))

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 42
    for col in range(2, len(roles) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_role_profile_template(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if not header or header[0].lower() != "competencia":
        return 0, ["La primera columna debe ser 'Competencia'."]

    role_by_name = {r.name: r for r in Role.objects.all()}
    roles = []
    errors = []

    for col_idx, role_name in enumerate(header[1:], start=2):
        role = role_by_name.get(role_name)
        if not role_name:
            continue
        if not role:
            errors.append(f"Rol '{role_name}' no existe en el sistema.")
            continue
        roles.append((col_idx, role))

    if errors:
        return 0, errors

    competency_by_name = {c.name: c for c in Competency.objects.all()}
    upserts = 0

    for row_idx in range(3, ws.max_row + 1):
        comp_name = ws.cell(row=row_idx, column=1).value
        if not comp_name:
            continue
        comp_name = str(comp_name).strip()
        competency = competency_by_name.get(comp_name)
        if not competency:
            errors.append(f"Competencia '{comp_name}' no existe.")
            continue

        for col_idx, role in roles:
            value = ws.cell(row=row_idx, column=col_idx).value
            if value in (None, ""):
                continue
            try:
                level = int(value)
            except (ValueError, TypeError):
                errors.append(f"Valor inválido en {comp_name} / {role.name}: '{value}'.")
                continue
            if level < 1:
                errors.append(f"Nivel inválido en {comp_name} / {role.name}: {level}.")
                continue

            RoleCompetencyRequirement.objects.update_or_create(
                role=role,
                competency=competency,
                defaults={"required_level": level, "weight": 1},
            )
            upserts += 1

    return upserts, errors
