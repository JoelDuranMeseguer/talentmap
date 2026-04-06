import io

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from competencies.models import Competency, RoleCompetencyRequirement
from people.models import Department, Role


class RoleProfileConfigTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="IT")
        self.role = Role.objects.create(name="Developer", department=self.dept)
        self.hr = User.objects.create_superuser("hr", "hr@test.com", "pass")
        self.comp = Competency.objects.create(name="Customer Centric")

    def test_download_template_contains_role_column(self):
        self.client.force_login(self.hr)
        response = self.client.get(reverse("download_role_profile_template"))
        self.assertEqual(response.status_code, 200)

        wb = load_workbook(io.BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        self.assertIn("Developer", headers)

    def test_manual_save_role_profile(self):
        self.client.force_login(self.hr)
        response = self.client.post(
            reverse("role_profile_config"),
            {
                "save_profile": "1",
                "role": self.role.id,
                f"comp_{self.comp.id}": "2",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        req = RoleCompetencyRequirement.objects.get(role=self.role, competency=self.comp)
        self.assertEqual(req.required_level, 2)

    def test_import_role_profile_excel(self):
        self.client.force_login(self.hr)
        template_resp = self.client.get(reverse("download_role_profile_template"))

        wb = load_workbook(io.BytesIO(template_resp.content))
        ws = wb.active

        target_row = None
        for row_idx in range(3, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == "Customer Centric":
                target_row = row_idx
                break
        self.assertIsNotNone(target_row)

        role_col = None
        for col_idx in range(2, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == "Developer":
                role_col = col_idx
                break
        self.assertIsNotNone(role_col)

        ws.cell(row=target_row, column=role_col, value=3)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        response = self.client.post(
            reverse("role_profile_config"),
            {
                "import_profile": "1",
                "excel_file": SimpleUploadedFile(
                    "perfil.xlsx",
                    out.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        req = RoleCompetencyRequirement.objects.get(role=self.role, competency=self.comp)
        self.assertEqual(req.required_level, 3)
