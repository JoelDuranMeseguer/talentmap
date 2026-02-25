"""Excel import/export for users and employees."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import Employee, Department, Role


def build_sample_excel():
    """
    Builds an Excel file with:
    - Sheet 1: Template with headers (first_name, last_name, email, department, role, manager_email)
    - Sheet 2: Current users ordered by department, role
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Plantilla_Importar"

    headers = ["nombre", "apellido", "email", "departamento", "rol", "manager_email", "manager_nombre", "manager_apellido"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    ws1.cell(row=2, column=1, value="Ejemplo")
    ws1.cell(row=2, column=2, value="Apellido")
    ws1.cell(row=2, column=3, value="ejemplo@empresa.com")
    ws1.cell(row=2, column=4, value="IT")
    ws1.cell(row=2, column=5, value="Developer")
    ws1.cell(row=2, column=6, value="jefe@empresa.com")
    ws1.cell(row=2, column=7, value="(opcional: si manager está en este Excel)")
    ws1.cell(row=2, column=8, value="")
    for col in range(1, 9):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    ws2 = wb.create_sheet("Usuarios_Actuales", 1)
    ws2.append(["nombre", "apellido", "email", "departamento", "rol", "manager"])
    for col in range(1, 7):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True)

    emps = (
        Employee.objects.filter(active=True)
        .select_related("user", "department", "role", "manager__user")
        .order_by("department__name", "role__name", "user__last_name", "user__first_name")
    )
    for emp in emps:
        manager_str = ""
        if emp.manager and emp.manager.user:
            manager_str = emp.manager.user.email or f"{emp.manager.user.first_name} {emp.manager.user.last_name}".strip()
        ws2.append([
            emp.user.first_name or "",
            emp.user.last_name or "",
            emp.user.email or "(sin email)",
            emp.department.name,
            emp.role.name,
            manager_str,
        ])
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    ws3 = wb.create_sheet("Departamentos", 2)
    ws3.append(["departamento"])
    ws3.cell(row=1, column=1).font = Font(bold=True)
    for d in Department.objects.all().order_by("name"):
        ws3.append([d.name])
    ws3.column_dimensions["A"].width = 25

    ws4 = wb.create_sheet("Roles", 3)
    ws4.append(["departamento", "rol"])
    ws4.cell(row=1, column=1).font = Font(bold=True)
    ws4.cell(row=1, column=2).font = Font(bold=True)
    for r in Role.objects.select_related("department").order_by("department__name", "name"):
        ws4.append([r.department.name, r.name])
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def parse_excel_import(file):
    """
    Parses uploaded Excel and returns list of dicts for import.
    Expected columns: nombre, apellido, email, departamento, rol, manager_email
    """
    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return [], ["El archivo está vacío."]

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    idx = {}
    required = ["nombre", "apellido", "email", "departamento", "rol"]
    for name in required:
        try:
            idx[name] = header.index(name)
        except ValueError:
            return [], [f"Falta la columna '{name}' en el archivo."]
    idx["email"] = header.index("email") if "email" in header else -1
    idx["manager_email"] = header.index("manager_email") if "manager_email" in header else -1
    idx["manager_nombre"] = header.index("manager_nombre") if "manager_nombre" in header else -1
    idx["manager_apellido"] = header.index("manager_apellido") if "manager_apellido" in header else -1

    depts = {d.name: d for d in Department.objects.all()}
    roles_by_dept = {}
    for r in Role.objects.select_related("department"):
        key = (r.department.name, r.name)
        roles_by_dept[key] = r
    emp_by_email = {e.user.email.lower(): e for e in Employee.objects.filter(active=True) if e.user.email}
    all_excel_emails = set()
    for row in rows[1:]:
        if not row or idx["email"] < 0 or idx["email"] >= len(row):
            continue
        raw_email = str(row[idx["email"]] or "").strip().lower()
        if raw_email:
            all_excel_emails.add(raw_email)

    excel_emails = set()

    result = []
    errors = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(v for v in row):
            continue
        def _s(val):
            return (str(val or "").strip()).strip()

        def _cell(k):
            i = idx.get(k, -1)
            if i < 0 or i >= len(row):
                return ""
            return _s(row[i])

        nombre = _cell("nombre")
        apellido = _cell("apellido")
        email = _cell("email")
        dept_name = _cell("departamento")
        rol_name = _cell("rol")
        manager_email = _cell("manager_email").lower()
        manager_nombre = _cell("manager_nombre")
        manager_apellido = _cell("manager_apellido")

        if not nombre or not apellido:
            errors.append(f"Fila {i}: nombre y apellido son obligatorios.")
            continue
        if not email:
            errors.append(f"Fila {i}: email es obligatorio.")
            continue
        email = email.lower()
        if email in excel_emails:
            errors.append(f"Fila {i}: email '{email}' está duplicado en el archivo.")
            continue
        excel_emails.add(email)
        if not dept_name:
            errors.append(f"Fila {i}: departamento es obligatorio.")
            continue
        if not rol_name:
            errors.append(f"Fila {i}: rol es obligatorio.")
            continue

        dept = depts.get(dept_name)
        if not dept:
            errors.append(f"Fila {i}: departamento '{dept_name}' no existe. Crear en Config.")
            continue
        role = roles_by_dept.get((dept_name, rol_name))
        if not role:
            errors.append(f"Fila {i}: rol '{rol_name}' en departamento '{dept_name}' no existe.")
            continue

        manager = None
        manager_email_pending = ""
        if manager_email:
            manager = emp_by_email.get(manager_email.lower())
            if manager:
                pass
            elif manager_email in all_excel_emails:
                manager_email_pending = manager_email
            else:
                errors.append(
                    f"Fila {i}: manager_email '{manager_email}' no coincide con ningún empleado activo y no existe en este Excel."
                )
                continue
        elif manager_nombre and manager_apellido:
            result.append({
                "first_name": nombre,
                "last_name": apellido,
                "email": email,
                "department": dept,
                "role": role,
                "manager": None,
                "manager_email_pending": "",
                "manager_nombre": manager_nombre,
                "manager_apellido": manager_apellido,
                "row": i,
            })
            continue

        result.append({
            "first_name": nombre,
            "last_name": apellido,
            "email": email,
            "department": dept,
            "role": role,
            "manager": manager,
            "manager_email_pending": manager_email_pending,
            "manager_nombre": "",
            "manager_apellido": "",
            "row": i,
        })

    return result, errors
